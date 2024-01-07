#!/usr/bin/python3
# Encoding: utf-8 -*-
# @Time : 2023/12/16 22:54
# @Author : qifan
# @Email : qifan.wang@westwell-lab.com
# @File : plain_dos_page.py
# @Project : PyQtDemo

import os
import time
import pandas as pd
from math import isnan
from loguru import logger
from openpyxl import workbook, styles
from dateutil import parser
from tkinter import Tk
from tkinter.filedialog import askopenfilename

logger.remove()  # remove log to std
logger.add('record.log')


def get_struct_from_input():
    logger.debug(f"结算单输入信息收集，默认拉板对回板为一对多（{time.ctime()}）")

    choice = input("1. 手动输入\n2. 读取Excel文件\n").strip()
    while choice not in ['1', '2']:
        print("输入有误请重新输入！")
        choice = input("1. 手动输入\n2. 读取Excel文件\n").strip()
        continue

    result_structure, from_list, to_list = list(), list(), list()
    if choice == '1':
        unit_price = None
        while True:
            try:
                unit_price = input("请输入拉板/回板单价，默认全程使用该价格进行计算！\n")
                unit_price = float(unit_price)
                break
            except ValueError:
                print(f"输入有误：{unit_price}，请重新输入！")

        idx = 1
        while True:
            try:
                from_date = input(f"请输入第{idx}次【拉板】日期（格式：YYYY/MM/DD或DD/MM/YY），键入字母“e”结束输入！\n")
                if from_date.strip().lower() == 'e':
                    break
                elif from_date.strip() == '':
                    print("输入有误请重新输入！")
                    continue
                from_date = parser.parse(from_date).date()
                from_amount = int(input("请输入该次【拉板】块数\n"))
                from_list.append([from_date, from_amount])

                idx += 1
            except Exception:
                print("日期/数字格式输入有误，请检查后重新输入！")
        logger.debug(f"输入了{idx - 1}个拉板日期，合计块数{sum(i[-1] for i in from_list)}")

        idx = 1
        while True:
            try:
                to_date = input(f"请输入第{idx}次【回板】日期（格式：YYYY/MM/DD或DD/MM/YY），键入字母“e”结束输入！\n")
                if to_date.strip().lower() == 'e':
                    break
                elif to_date.strip() == '':
                    print("输入有误请重新输入！")
                    continue
                to_date = parser.parse(to_date).date()
                to_amount = int(input("请输入该次【回板】块数\n"))
                to_list.append([to_date, to_amount])

                idx += 1
            except Exception:
                print("日期/数字格式输入有误，请检查后重新输入！")
        logger.debug(f"输入了{idx - 1}个回板日期，合计块数{sum(i[-1] for i in to_list)}")

    else:
        print("使用此方式需excel文件严格按照五列【拉板日期、拉板块数、回板日期、回板块数、单价】进行值填充！")
        time.sleep(0.75)
        Tk().withdraw()  # https://stackoverflow.com/questions/3579568/choosing-a-file-in-python-with-simple-dialog
        filename = askopenfilename()
        if not filename:
            print("Not selected any file. Gonna populate nothing into target excel file!")
            return 0, []
        logger.debug(f"opened file {filename}")
        df = pd.DataFrame(pd.read_excel(io=filename))
        dfv = df.values.T.tolist()

        from_date_l = [parser.parse(i).date() for i in dfv[0] if type(i) is not float]
        from_count_l = [i for i in dfv[1] if not isnan(float(i))]
        to_date_l = [parser.parse(i).date() for i in dfv[2] if type(i) is not float]
        to_count_l = [i for i in dfv[3] if not isnan(float(i))]
        unit_price = dfv[4][0]

        for x in zip(*[from_date_l, from_count_l]):
            from_list.append(list(x))
        for x in zip(*[to_date_l, to_count_l]):
            to_list.append(list(x))

    for from_item in from_list:
        sublist = []
        from_val = from_item[-1]
        while from_val != 0:
            if to_list and to_list[0][-1] <= from_val:
                sublist.append(to_list.pop(0))
                from_val -= sublist[-1][-1]
            else:
                temp_val = to_list[0].copy()
                temp_val[-1] = from_val
                sublist.append(temp_val)
                to_list[0][-1] -= from_val
                from_val = 0

        res = from_item.copy()
        res.append(sublist)
        result_structure.append(res)

    logger.debug(f"单价：{unit_price}")
    logger.debug(f"结构：{result_structure}")

    return unit_price, result_structure


def form_xlsx_file(price, structure):
    workbook_ = workbook.Workbook()
    sheet_ = workbook_.active

    # region manipulate the cells
    center_alignment = styles.Alignment(horizontal='center', vertical='center')

    # header part
    sheet_.merge_cells('C3:D3')
    sheet_['C3'], sheet_['C4'], sheet_['D4'] = "拉板", "日期", "块数"
    sheet_['C3'].alignment = center_alignment

    sheet_.merge_cells('E3:F3')
    sheet_['E3'], sheet_['E4'], sheet_['F4'] = "回板", "日期", "块数"
    sheet_['E3'].alignment = center_alignment

    sheet_.merge_cells('G3:G4')
    sheet_.merge_cells('H3:H4')
    sheet_.merge_cells('I3:I4')
    sheet_['G3'], sheet_['H3'], sheet_['I3'] = '天数', '单价', '总价'
    sheet_['G3'].alignment, sheet_['H3'].alignment, sheet_['I3'].alignment = \
        center_alignment, center_alignment, center_alignment

    # body part
    start_row_, start_column = 5, 3
    for item in structure:
        depth_ = len(item[-1])
        if depth_ == 1:
            frm_date_cell = sheet_.cell(row=start_row_, column=start_column, value=item[0])  # 拉板日期
            sheet_.cell(row=start_row_, column=start_column + 1, value=item[1])  # 拉板块数
            to_date_cell = sheet_.cell(row=start_row_, column=start_column + 2, value=item[-1][0][0])  # 回板日期
            to_amnt_cell = sheet_.cell(row=start_row_, column=start_column + 3, value=item[-1][0][1])  # 回板块数
            day_cnt_cell = sheet_.cell(row=start_row_, column=start_column + 4,
                                       value=f"={to_date_cell.coordinate}-{frm_date_cell.coordinate}")  # 天数
            uni_prc_cell = sheet_.cell(row=start_row_, column=start_column + 5, value=price)  # 单价
            sheet_.cell(row=start_row_, column=start_column + 6,
                        value=f'={to_amnt_cell.coordinate}*{day_cnt_cell.coordinate}*{uni_prc_cell.coordinate}')  # 总价

        else:
            sheet_.merge_cells(start_row=start_row_, start_column=start_column, end_row=start_row_ + depth_ - 1,
                               end_column=start_column)
            sheet_.merge_cells(start_row=start_row_, start_column=start_column + 1, end_row=start_row_ + depth_ - 1,
                               end_column=start_column + 1)
            frm_date_cell = sheet_.cell(row=start_row_, column=start_column, value=item[0])  # 拉板日期
            frm_date_cell.alignment = center_alignment
            sheet_.cell(row=start_row_, column=start_column + 1, value=item[1]).alignment = center_alignment  # 拉板块数

            for idx, sub_item in enumerate(item[-1]):
                row = start_row_ + idx
                to_date_cell = sheet_.cell(row=row, column=start_column + 2, value=sub_item[0])  # 回板日期
                to_amnt_cell = sheet_.cell(row=row, column=start_column + 3, value=sub_item[1])  # 回板块数
                day_cnt_cell = sheet_.cell(row=row, column=start_column + 4,
                                           value=f"={to_date_cell.coordinate}-{frm_date_cell.coordinate}")  # 天数
                uni_prc_cell = sheet_.cell(row=row, column=start_column + 5, value=price)  # 单价
                sheet_.cell(row=row, column=start_column + 6,
                            value=f"={to_amnt_cell.coordinate}*{day_cnt_cell.coordinate}*{uni_prc_cell.coordinate}")  # 总价

        start_row_ += (depth_ + 1)

    # sum
    sheet_.cell(row=start_row_, column=start_column - 1, value="总计")
    sheet_.cell(row=start_row_, column=start_column + 6, value=f"=SUM(I5:I{start_row_ - 1})")
    # endregion manipulate the cells

    filename_ = os.path.join(os.path.join(os.environ['USERPROFILE'], 'Desktop'), '年度钢板结算单.xlsx')
    logger.info(f"the file is saved to {filename_}\n")

    workbook_.save(filename=filename_)
    return filename_


def main():
    price, structure = get_struct_from_input()
    # price, structure = 2.5, [[datetime.date(2023, 1, 1), 20, [[datetime.date(2023, 2, 1), 10], [datetime.date(2023, 2, 5), 5], [datetime.date(2023, 3, 3), 5]]], [datetime.date(2023, 2, 1), 15, [[datetime.date(2023, 5, 1), 10], [datetime.date(2023, 5, 10), 5]]], [datetime.date(2023, 3, 3), 5, [[datetime.date(2024, 5, 5), 5]]]]
    filename = form_xlsx_file(price, structure)
    print(f"已将输出信息保存到{filename}中！\n")
    os.system("pause")


if __name__ == '__main__':
    # From https://chat.openai.com/c/34221f1f-8901-42dc-9ff7-0929658ef3fb
    # pyinstaller plain_dos_page.py --onefile -n SillyBee
    main()
